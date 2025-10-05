import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By


# 1. Load the workbook
filename=r"C:\Users\Admin\OneDrive\Desktop/Tejasdata.xlsx"
my_workbook=openpyxl.load_workbook(filename)

#2. get the sheet

my_sheet=my_workbook["Sheet1"]

#3. get the no of rows and columns

rows=my_sheet.max_row
columns=my_sheet.max_column
print(rows)
print(columns)
#4. get the data    

    
    
def test_orangehrm_login(username,password,url,i):
        # 1. Lauching Chrome Browser with desired capabilities
        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach", True)
        options.add_argument("start-maximized")
        driver = webdriver.Chrome(options=options)
        driver.implicitly_wait(5)
        
        # 2. Navigating to OrangeHRM Login Page
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        # 3. Enter Username
        username_txt_bx = driver.find_element(By.XPATH, "//input[@name='username']")
        username_txt_bx.send_keys(username)
    
    #4. Enter Password
        passsword_txt_bx = driver.find_element(By.XPATH, "//input[@name='password']")
        passsword_txt_bx.send_keys(password)
        
        # 5. Click Login Button
        login_btn = driver.find_element(By.XPATH, "//button[@type='submit']")
        login_btn.click()
        
        actual_url = driver.current_url
        expected_url="/dashboard/index"
        
        if expected_url in actual_url:
            print("Testcase passed ")
            my_sheet.cell(i,4).value="Testcase passed"
            my_workbook.save(filename)

        elif url in actual_url:
            print("test case passed")
            my_sheet.cell(i,4).value="Testcase passed"
            my_workbook.save(filename)


        else:
            print("Testcase failed")
            my_sheet.cell(i,4).value="Testcase failed"
            my_workbook.save(filename)

            
            
        
        
    #loop through the xecel data and call the login method
    
for i in range(2, rows + 1):
    username = my_sheet.cell(i, 1).value
    password = my_sheet.cell(i, 2).value
    url = my_sheet.cell(i, 3).value
    
    test_orangehrm_login(username, password,url,i)
        
    
    
    