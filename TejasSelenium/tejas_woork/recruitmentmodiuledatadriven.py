'''
Created on 20-Sept-2025

@author: Admin
'''


from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By


def test_login_to_orange_hrm(username,password,url,i):
    #1.lauching the browser
        options=webdriver.ChromeOptions()
        options.add_experimental_option("detach",True)
        options.add_argument("start-maximized")
        driver=webdriver.Chrome(options=options)
        driver.implicitly_wait(5)
        
        #2.navigate to url
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        
        if not username or not password:
            
            print("invalid input please enter the correct input")
            url=driver.current_url
            if "/auth/login" in url:
                my_sheet.cell(i, 4).value = "Testcase passed"
                my_workbook.save(filename)
                driver.quit()
                return
        
                 
        #3.enter username
            username_txt_bx=driver.find_element(By.XPATH,"//input[@name='username']")
            username_txt_bx.send_keys(username)
        
        
        #4.enter password
            password_txt_bx=driver.find_element(By.XPATH,"//input[@name='password']")
            password_txt_bx.send_keys(password)
        
        
            
        #5.click login button
        login_btn=driver.find_element(By.XPATH,"//button[@type='submit']")
        login_btn.click()

        
        actual_url=driver.current_url
        if url in actual_url:
            print("test case passed")
            my_sheet.cell(i,4).value="test case passed"
            my_workbook.save(filename)
        else:
            print("testcase failed")
            my_sheet.cell(i,4).value="testcase failed"
            my_workbook.save(filename)
            
    
# 1. Load the workbook
filename="C:\\Users\\Admin\\OneDrive\\Desktop\\Narayanaswamy.xlsx"

my_workbook=openpyxl.load_workbook(filename)

#2.get the sheet
my_sheet=my_workbook["Sheet1"]
#3.get the no of rows and columns

rows=my_sheet.max_row
columns=my_sheet.max_column

print(rows)
print(columns)

for i in range(2,rows+1):
    username=my_sheet.cell(i,1).value
    password=my_sheet.cell(i,2).value
    url=my_sheet.cell(i,3).value
    print(username,password,url)

    test_login_to_orange_hrm(username, password, url, i)
    
   